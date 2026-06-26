#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
一键导出问卷收集数据为 Excel 表格。

从 Supabase 拉取全部 users / submissions，按「一个用户一条数据」归并
（优先已提交，否则取最新草稿），解码选项值为中文，输出带状态色标的 xlsx。

用法（必须在 conda base 运行，禁止本地/系统 python）：
    /opt/miniconda3/bin/python scripts/export_survey.py
或：
    npm run export

凭据从项目根目录 .env 读取：VITE_SUPABASE_URL / VITE_SUPABASE_ANON_KEY。
输出文件：survey_data_by_user_<YYYY-MM-DD>.xlsx（已被 .gitignore 忽略，含联系方式，禁止提交）。
"""
import argparse
import sys
from collections import Counter
from datetime import date
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent


# ---------------------------------------------------------------- 凭据
def load_env() -> tuple[str, str]:
    """从根目录 .env 读取 Supabase URL / anon key（也兼容裸 SUPABASE_*）。"""
    env = {}
    env_path = ROOT / ".env"
    if env_path.exists():
        for line in env_path.read_text().splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip().strip('"').strip("'")
    url = env.get("VITE_SUPABASE_URL") or env.get("SUPABASE_URL")
    key = env.get("VITE_SUPABASE_ANON_KEY") or env.get("SUPABASE_ANON_KEY")
    if not url or not key:
        sys.exit("[ERROR] 未在 .env 找到 VITE_SUPABASE_URL / VITE_SUPABASE_ANON_KEY")
    return url.rstrip("/"), key


def fetch(url: str, key: str, path: str) -> list[dict]:
    r = requests.get(
        f"{url}/rest/v1/{path}",
        headers={"apikey": key, "Authorization": f"Bearer {key}"},
        timeout=60,
    )
    r.raise_for_status()
    return r.json()


# ---------------------------------------------------------------- 选项值 -> 中文映射
M = {
    "q2": {"1": "愿意提供并公开", "2": "愿意提供不公开", "3": "不愿意提供"},
    "q4": {"math_2p2": "数学与应用数学(2+2)", "math_4p0": "数学与应用数学(4+0)", "stats": "统计"},
    "q5": {"phd": "博士", "master": "硕士"},
    "q6": {"statistics": "Statistics", "data_science": "Data Science", "ml_ai": "ML/AI",
           "fin_math": "金融数学/金融统计", "pure_math": "纯数", "applied_math": "应用数学/计算数学",
           "biostat": "生物统计", "analytics": "Analytics/BA", "or": "运筹", "business": "经管类", "other": "其他"},
    "q7": {"uk": "英国", "us": "美国", "hk": "香港", "sg": "新加坡", "au": "澳大利亚", "eu": "欧陆", "other": "其他"},
    "q8": {"return_work": "回国就业", "overseas_work": "留外就业", "further_study": "继续深造", "other": "其他"},
    "q9_status": {"decided": "去向已定", "undecided_willing": "未定·愿更新", "undecided_not": "未定·不愿更新"},
    "q14": {"none": "无未出结果项目", "willing": "有·愿更新", "not_willing": "有·不愿更新"},
    "q15": {"willing": "愿意", "not_willing": "不愿意"},
    "q21": {"full": "全包", "half": "半包", "diy": "DIY", "other": "其他"},
    "q22": {"yes": "有", "no": "无"},
    "q24": {"yes": "有", "no": "无"},
}
ENTRY = {"school": "学校", "project": "项目", "submitTime": "提交时间", "receiveTime": "收到时间",
         "cond": "Cond", "scholarship": "奖学金", "note": "备注", "time": "时间", "institution": "机构/地点",
         "title": "项目/职位", "advisor": "导师", "duration": "时长", "content": "内容", "output": "产出",
         "company": "企业/地点"}

FIELDS = [
    ("q1", "Q1 姓名/代号"), ("q2", "Q2 联系意愿"), ("q3_email", "Q3 邮箱"), ("q3_wechat", "Q3 微信"),
    ("q3_other", "Q3 其他联系"), ("q4", "Q4 本科专业"), ("q5", "Q5 申请学位"), ("q6", "Q6 方向偏好"),
    ("q6_other_text", "Q6 其他方向"), ("q7", "Q7 申请地区"), ("q8", "Q8 申请目的"), ("q9_status", "Q9 去向状态"),
    ("q9_text", "Q9 去向详情"), ("q10_gpa_pct", "Q10 均分"), ("q10_gpa_4", "Q10 GPA"), ("q10_language", "Q10 语言"),
    ("q10_gre", "Q10 GRE/GMAT"), ("q11", "Q11 Admission"), ("q12", "Q12 Waitlist"), ("q13", "Q13 Reject"),
    ("q14", "Q14 未出结果"), ("q15", "Q15 愿提供经历"), ("q16", "Q16 科研/项目"), ("q17", "Q17 实习"),
    ("q19", "Q19 推荐信"), ("q20", "Q20 荣誉奖项"), ("q21", "Q21 申请方式"), ("q22", "Q22 中介分享意愿"),
    ("q23", "Q23 中介分享"), ("q24", "Q24 DIY分享意愿"), ("q25", "Q25 DIY分享"), ("q26", "Q26 申请心得"),
]
STRUCT = {"q11", "q12", "q13", "q16", "q17"}


def fmt(key: str, val) -> str:
    """把 draft 里的原始值解码成可读中文。"""
    if val is None or val == "" or val == []:
        return ""
    if key in STRUCT:  # 矩阵/列表型：每条目「字段:值」分号连接，多条目换行
        if isinstance(val, str):
            return val
        if isinstance(val, list):
            return "\n".join("；".join(f"{ENTRY.get(k, k)}:{v}" for k, v in e.items() if v) for e in val)
        return str(val)
    if isinstance(val, list):  # 多选
        return "、".join(M.get(key, {}).get(v, str(v)) for v in val)
    return M.get(key, {}).get(val, str(val))


def filled_count(d: dict) -> int:
    return sum(1 for k, _ in FIELDS if d.get(k) not in (None, "", []))


# ---------------------------------------------------------------- 主流程
def main() -> None:
    ap = argparse.ArgumentParser(description="一键导出问卷数据为 Excel")
    ap.add_argument("-o", "--output", help="输出文件路径（默认 ./survey_data_by_user_<日期>.xlsx）")
    args = ap.parse_args()

    url, key = load_env()
    print(f"[INFO] 拉取 Supabase 数据: {url}")
    subs = fetch(url, key, "submissions?select=id,user_id,status,submitted_at,updated_at,draft&order=updated_at.desc")
    users = {u["id"]: u for u in fetch(url, key, "users?select=id,identifier,identifier_type,created_at")}
    print(f"[INFO] submissions={len(subs)}  users={len(users)}")

    # 按用户归并：一个用户一条（优先已提交，再按提交/更新时间最新）。无 submission 的用户自动忽略。
    def rank(s):
        return (1 if s["status"] == "submitted" else 0, s.get("submitted_at") or "", s.get("updated_at") or "")
    best = {}
    for s in subs:
        uid = s["user_id"]
        if uid not in best or rank(s) > rank(best[uid]):
            best[uid] = s

    # 排序：已提交在前（按提交时间），草稿在后（按已填字段数降序）
    def sortkey(s):
        d = s.get("draft") or {}
        return (0, s.get("submitted_at") or "") if s["status"] == "submitted" else (1, -filled_count(d))
    rows = sorted(best.values(), key=sortkey)

    # ---- 写 Excel ----
    wb = Workbook()
    ws = wb.active
    ws.title = "问卷数据"
    HEAD = ["#", "状态", "已填字段数", "身份(手机/邮箱)", "提交时间", "更新时间"] + [lbl for _, lbl in FIELDS]
    ws.append(HEAD)

    green = PatternFill("solid", fgColor="C6EFCE")
    amber = PatternFill("solid", fgColor="FFEB9C")
    headfill = PatternFill("solid", fgColor="305496")
    headfont = Font(color="FFFFFF", bold=True, size=10)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in ws[1]:
        c.fill = headfill
        c.font = headfont
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    for i, s in enumerate(rows, 1):
        d = s.get("draft") or {}
        u = users.get(s["user_id"], {})
        base = [i, "已提交" if s["status"] == "submitted" else "草稿", filled_count(d),
                u.get("identifier", ""), (s.get("submitted_at") or "")[:19].replace("T", " "),
                (s.get("updated_at") or "")[:19].replace("T", " ")]
        ws.append(base + [fmt(k, d.get(k)) for k, _ in FIELDS])
        ws.cell(row=i + 1, column=2).fill = green if s["status"] == "submitted" else amber

    for col, w in (("A", 4), ("B", 8), ("C", 10), ("D", 22), ("E", 18), ("F", 18)):
        ws.column_dimensions[col].width = w
    for idx in range(7, 7 + len(FIELDS)):
        ws.column_dimensions[get_column_letter(idx)].width = 22
    for r in ws.iter_rows(min_row=2):
        for c in r:
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = border
    ws.freeze_panes = "G2"
    ws.row_dimensions[1].height = 30

    out = Path(args.output) if args.output else ROOT / f"survey_data_by_user_{date.today().isoformat()}.xlsx"
    wb.save(out)

    submitted = sum(1 for s in rows if s["status"] == "submitted")
    print(f"[SUCCESS] 已生成: {out}")
    print(f"          共 {len(rows)} 位用户  |  已提交 {submitted}  |  仅草稿 {len(rows) - submitted}"
          f"  |  另有 {len(users) - len(best)} 位登录但无任何提交（已忽略）")


if __name__ == "__main__":
    main()
